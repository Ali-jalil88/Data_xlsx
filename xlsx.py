from openpyxl import load_workbook

"""
In this code:
1. The `openpyxl` library is imported to handle Excel files.
2. The path to the Excel file `MOCK_DATA.xlsx` is defined.
3. The Excel file is loaded using `load_workbook`.
4. The names of the sheets in the file are printed.
5. The active sheet is selected.
6. Each row in the active sheet is printed.
7. New rows (`new_row` and `new_row2`) containing the data to be added are defined.
8. The new rows are appended to the active sheet using the `append` method.
9. The changes are saved to the Excel file using the `save` method.
10. The Excel file is closed.
"""

# Meine Frage für Herr Ahmed warum new_row kopieret wieder 2Mal, Wenn drücke Run.

# Define the path to the Excel file
file_path = "MOCK_DATA.xlsx"

# Load the Excel file
workbook = load_workbook(filename = file_path)

# Print the names of the sheets in the Excel file
print(workbook.sheetnames)

# Select the active sheet
sheet = workbook.active

# Print each row in the active sheet
for row in sheet.iter_rows(values_only= True):
    print(row)

# Define new rows to be added
new_row = (11, 'Hassan', 'Ahmed', 'Hassan_H5854545@sphinn.com', 'Male', '123.142.131.76')
new_row2 = (12, 'Ahmed', 'Hassan', 'Ahmed_Hassan16547@sphinn.com', 'Male', '123.142.131.76')

# Append the new rows to the active sheet
sheet.append(new_row)
sheet.append(new_row2)

# Save the changes to the Excel file
workbook.save(filename=file_path)

# Close the Excel file
workbook.close()